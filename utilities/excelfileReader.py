import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_col_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path, sheet_name, num_row, num_col):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=num_row, column=num_col).value

def set_cell_data(path, sheet_name, num_row, num_col, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=num_row, column=num_col).value = data

def get_excel_data(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_num_rows = sheet.max_row
    total_num_cols = sheet.max_column

    final_list = []
    for r in range(1, total_num_rows + 1):
        row_list = []
        for c in range(1, total_num_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    
    return final_list
